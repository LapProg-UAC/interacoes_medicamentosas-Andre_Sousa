import random
import csv
import openpyxl
from typing import List, Dict, Tuple, Any, TextIO

def read_meds(filename: str) -> List[str]:
    """
    Lê um ficheiro de texto e extrai os nomes dos medicamentos para uma lista.

    Parameters
    ----------
    filename : str
        O caminho para o ficheiro de texto a ser lido no sistema de ficheiros.

    Returns
    -------
    list of str
        Uma lista contendo os nomes dos medicamentos limpos e sem espaços em branco adicionais.

    Raises
    ------
    FileNotFoundError
        Se o ficheiro de texto especificado não for encontrado.
    OSError
        Se ocorrer qualquer outro erro inesperado durante a abertura ou leitura do ficheiro.
    """
    clean_list: List[str] = []
    try:
        file_handler: TextIO
        with open(filename, 'r', encoding='utf-8') as file_handler:
            line: str
            for line in file_handler:
                clean_text: str = line.strip()
                if clean_text:
                    clean_list.append(clean_text)
        return clean_list
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Erro: O ficheiro '{filename}' não foi encontrado.") from exc
    except Exception as exc:
        raise OSError(f"Erro inesperado ao ler o ficheiro: {exc}") from exc

def gen_matrix(meds: List[str]) -> Dict[Tuple[str, str], int]:
    """
    Gera uma matriz de interações medicamentosas estruturada em dicionário e tuplos.

    Parameters
    ----------
    meds : list of str
        A lista com os nomes de todos os medicamentos extraídos.

    Returns
    -------
    dict
        Dicionário cujas chaves são tuplos de pares de medicamentos (estrutura imutável) 
        e os valores são o grau de interação (0 para o mesmo medicamento, 1 a 6 para distintos).
    """
    matrix: Dict[Tuple[str, str], int] = {}
    med_row: str
    med_col: str
    for med_row in meds:
        for med_col in meds:
            if med_row == med_col:
                matrix[(med_row, med_col)] = 0
            else:
                matrix[(med_row, med_col)] = random.randint(1, 6)
    return matrix

def export_excel(matrix: Dict[Tuple[str, str], int], meds: List[str], filename: str) -> None:
    """
    Exporta os dados da matriz de interações para uma folha de cálculo em formato Excel.

    Parameters
    ----------
    matrix : dict
        O dicionário de interações gerado com as combinações de todos os medicamentos.
    meds : list of str
        A lista com os nomes dos medicamentos a ser utilizada no cabeçalho e na primeira coluna.
    filename : str
        O nome ou diretório do ficheiro de destino onde o documento Excel será guardado.

    Raises
    ------
    TypeError
        Se a matriz submetida for nula ou não inicializada.
    OSError
        Se houver problemas de permissão, ficheiro corrompido ou falhas ao guardar no disco.
    """
    if matrix is None:
        raise TypeError("Erro Crítico: A matriz de dados não pode ser nula.")
    
    try:
        workbook: openpyxl.Workbook = openpyxl.Workbook()
        worksheet: Any = workbook.active
        
        header_row: List[str] = [""]
        med: str
        for med in meds:
            header_row.append(med)
        
        worksheet.append(header_row)
        
        row_med: str
        col_med: str
        for row_med in meds:
            new_row: List[Any] = [row_med]
            for col_med in meds:
                effect: int = matrix[(row_med, col_med)]
                new_row.append(effect)
            
            worksheet.append(new_row)
            
        workbook.save(filename)
    except Exception as exc:
        raise OSError(f"Erro grave ao tentar criar ou guardar o ficheiro Excel: {exc}") from exc

def convert_xlsx_to_csv(excel_path: str, csv_path: str) -> None:
    """
    Converte um ficheiro Excel para um ficheiro CSV rigorosamente formatado.

    A função utiliza a biblioteca openpyxl para carregar a folha ativa de um
    documento Excel e converte-o de imediato, escrevendo o conteúdo linha a
    linha num ficheiro de destino suportado nativamente pelo formato CSV.

    Parameters
    ----------
    excel_path : str
        Caminho exato no disco do ficheiro Excel de origem.
    csv_path : str
        Caminho do ficheiro CSV a ser criado como destino.

    Raises
    ------
    FileNotFoundError
        Se o ficheiro Excel indicado não for localizado no sistema.
    IOError
        Se ocorrer qualquer anomalia de input/output ou de permissões.
    """
    try:
        workbook: openpyxl.Workbook = openpyxl.load_workbook(excel_path)
        worksheet: Any = workbook.active
        
        csv_file: TextIO
        with open(csv_path, 'w', encoding='utf-8', newline='') as csv_file:
            csv_writer: Any = csv.writer(csv_file, delimiter=',')
            
            row: Tuple[Any, ...]
            for row in worksheet.iter_rows(values_only=True):
                row_data: List[Any] = list(row)
                csv_writer.writerow(row_data)

    except FileNotFoundError as exc:
        raise FileNotFoundError(f"Erro: O ficheiro Excel de origem '{excel_path}' não foi encontrado.") from exc
  